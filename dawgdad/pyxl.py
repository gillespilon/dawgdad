"""
openpyxl functions
"""

from pathlib import Path
from typing import IO
import time
import sys
import io

from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, \
    Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.workbook.workbook import Workbook
from dawgdad import report_summary, html_end
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import pandas as pd
import numpy as np


def autofit_column_width(
    *,
    ws: Worksheet,
    extra_width: int
) -> Worksheet:
    """
    Autofit all columns in a worksheet.

    Parameters
    ----------
    ws : Worksheet
        The worksheet in which to autofit all columns.
    extra_width : int
        An integer to add extra width so that the column edges are not flush.

    Returns
    -------
    ws : Worksheet
        The worksheet in which autofit was applied to all columns

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.autofit_column_width(
    ...     ws=ws,
    ...     extra_width=7
    ... )
    """
    for column in ws.columns:
        max_width = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:  # Necessary to avoid error on empty cells
                if len(str(cell.value)) > max_width:
                    max_width = len(cell.value)
            except Exception:
                pass
        adjusted_width = (max_width + extra_width)
        ws.column_dimensions[column_letter].width = adjusted_width
    return ws


def cell_fill_down(
    *,
    ws: Worksheet,
    min_row: int,
    max_row: int,
    min_col: int,
    max_col: int
) -> Worksheet:
    """
    Fill empty cell with the value from the cell above

    Parameters
    ----------
    ws : Worksheet
        The worksheet in which to change the case of column(s).
    min_row : int
        The first row in the range to change.
    max_row : int
        The last row in the range to change.
    min_col : int
        The first column in the range to change.
    max_col : int
        The last column in the range to change.

    Returns
    -------
    ws : Worksheet
        The worksheet in which cells were modified.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> fill_down_columns = ["col1", "col2", "col3"]
    >>> for column in fill_down_columns:
    ...     ws = dd.cell_fill_down(
    ...         ws=ws,
    ...         min_row=2,
    ...         max_row=ws.max_row,
    ...         min_col=1,
    ...         max_col=3
    ...     )
    """
    row_count = 0
    for row in ws.iter_rows(
        min_col=min_col,
        max_col=max_col
    ):
        for cell in row:
            if cell.value:
                row_count += 1
    if row_count > 0:
        for row in ws.iter_rows(
            min_row=min_row + 1,  # start one row below the 'start' row
            max_row=max_row,
            min_col=min_col,
            max_col=max_col
        ):
            for cell in row:
                if cell.value in [None, 'None', '']:
                    cell.value = ws[cell.row - 1][min_col - 1].value
    return ws


def cell_style(
    *,
    wb: Workbook,
    style_name: str = 'cell_style',
    font_name: str = 'Lucida Sans',
    font_size: int = 11,
    font_bold: bool = True,
    font_colour: str = '000000',
    horizontal_alignment: str = 'center',
    vertical_alignment: str = 'center',
    wrap_text: str | bool = None,
    fill_type: str | bool = 'solid',
    foreground_colour: str | bool = 'd9d9d9',
    border_style: str | bool = None,
    border_colour: str | bool = None,
    number_format: str | bool = None
) -> NamedStyle:
    """
    Define a cell style

    Parameters
    ----------
    wb : Workbook
        The workbook in which to define the cell style.
    style_name : str = 'cell_style'
        The name for the cell style.
    font_name : str = 'Lucida Sans'
        The font name for the style.
    font_size : int = 11
        The font size for the style.
    font_bold : bool = True
        A boolean or string to apply bold style.
    font_colour : str = 'ffffff'
        The string for the font colour.
    horizontal_alignment : str = 'center'
        The string for horizontal alignment.
    vertical_alignment : str = 'center'
        The string for vertical alignment.
    wrap_text : str | bool = None
        A boolean or string to wrap text.
    fill_type : str = 'solid'
        The string for the fill type.
    foreground_colour : str = 'd9d9d9'
        The string for the foreground colour.
    border_style : str | bool = None
        A boolean or string to apply a border.
    border_colour : str | bool = None
        A boolean or string to apply a border colour.
    number_format : str | bool = None
        A boolean or string to apply a number format.

    Returns
    -------
    row_style : NamedStyle
        The named style.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> red_cell_style = dd.cell_style(
    ...     wb=wb,
    ...     style_name='red_cell_style',
    ...     font_colour='ffffff',
    ...     foreground_colour='c00000'
    ... )
    >>> wb.add_named_style(red_cell_style) # doctest: +SKIP
    >>> for cell in ['C1', 'D1', 'E1']:
    ...     ws[cell].style = red_cell_style # doctest: +SKIP
    """
    cell_style = NamedStyle(name=style_name)
    cell_style.font = Font(
        name=font_name,
        size=font_size,
        bold=font_bold,
        color=font_colour
    )
    cell_style.alignment = Alignment(
        horizontal=horizontal_alignment,
        vertical=vertical_alignment,
        wrap_text=wrap_text
    )
    cell_style.fill = PatternFill(
        fill_type=fill_type,
        fgColor=foreground_colour
    )
    cell_style.border = Border(
        bottom=Side(
            border_style=border_style,
            color=border_colour
        )
    )
    cell_style.number_format = number_format
    wb.add_named_style(cell_style)
    return (wb, cell_style)


def change_case_worksheet_columns(
    *,
    ws: Worksheet,
    min_col: int,
    max_col: int,
    min_row: int,
    max_row: int,
    case: str = 'upper'
) -> Worksheet:
    """
    Change case for one or more worksheet columns.

    Parameters
    ----------
    ws : Worksheet
        The worksheet in which to change the case of column(s).
    min_col : int
        The first column in the range to change.
    max_col : int
        The last column in the range to change.
    min_row : int
        The first row in the range to change.
    max_row : int
        The last row in the range to change.
    case: str = 'upper'
        The case to change. Currently only upper or lower.

    Returns
    -------
    ws : Worksheet
        A worksheet from a workbook.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.change_case_worksheet_columns(
    ...     ws=ws,
    ...     min_col=4,
    ...     max_col=6,
    ...     min_row=1,
    ...     max_row=ws.max_row,
    ...     case='upper'
    ... )
    """
    for col in ws.iter_cols(
        min_col=min_col,
        max_col=max_col,
        min_row=min_row,
        max_row=max_row
    ):
        for cell in col:
            if case == 'upper':
                cell.value = str(cell.value).upper()
            elif case == 'lower':
                cell.value = str(cell.value).lower()
    return ws


def exit_script(
    *,
    original_stdout: IO[str],
    output_url: str
) -> None:
    """
    Exit from a script and complete the html file.

    Parameters
    ----------
    original_stdout : IO[str]
        The original stdout.
    output_url : str
        The output url.

    Example
    -------
    >>> import dawgdad as dd
    >>> output_url = '../tests/my_html_file.html'
    >>> original_stdout = dd.html_begin(output_url=output_url)
    >>> dd.exit_script(
    ...     original_stdout=original_stdout,
    ...     output_url=output_url
    ... ) # doctest: +SKIP
    """
    html_end(
        original_stdout=original_stdout,
        output_url=output_url
    )
    sys.exit()


def list_duplicate_worksheet_rows(
    *,
    ws: Worksheet,
    columns_to_ignore: list[int] = None
) -> list[int]:
    """
    Find duplicate rows in a worksheet.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    columns_to_ignore : list[int] = None
        A list of column numbers to not use in determining duplicate rows.

    Returns
    -------
    duplicate_rows : list[int]
        A list of duplicate row numbers.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> duplicate_rows = dd.list_duplicate_worksheet_rows(ws=ws)
    >>> ws = dd.remove_worksheet_rows(
    ...     ws=ws,
    ...     rows_to_remove=duplicate_rows
    ... )
    """
    duplicate_rows = []
    unique_rows = []
    for row in ws.rows:
        working_list = []
        for cell in row:
            if cell.column not in columns_to_ignore:
                working_list.append(cell.value)
        if working_list not in unique_rows:
            unique_rows.append(working_list)
        else:
            duplicate_rows.append(cell.row)
    return duplicate_rows


def list_empty_and_nan_worksheet_rows(
    *,
    ws: Worksheet,
    min_row: int
) -> list[int]:
    """
    Create list of row numbers of blank worksheet rows.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    min_row : int
        Start row for iteration.

    Returns
    -------
    blank_rows : list[int]
        List of row numbers.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> sheetname = "sheetname"
    >>> ws = wb[sheetname] # doctest: +SKIP
    >>> blank_rows = dd.list_nan_worksheet_rows(
    ...     ws=ws,
    ...     min_row=2
    ... ) # doctest: +SKIP
    """
    blank_rows = []
    for row in ws.iter_rows(min_row=min_row):
        onerow = [cell.value for cell in row]
        if all(item in [
            None, 'None', 'NONE', '', np.nan
        ] for item in onerow):
            blank_rows.append(row[0].row)
    return blank_rows


def list_empty_except_nan_worksheet_rows(
    *,
    ws: Worksheet,
    min_row: int
) -> list[int]:
    """
    Create list of row numbers of empty worksheet rows, except those
    with np.nan.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    min_row : int
        Start row for iteration.

    Returns
    -------
    empty_rows : list[int]
        List of row numbers.

    Example
    -------
    Remove empty rows starting from row 2.
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> empty_rows = dd.list_empty_except_nan_worksheet_rows(
    ...     ws=ws,
    ...     min_row=2
    ... )
    """
    empty_rows = []
    for row in ws.iter_rows(min_row=min_row):
        onerow = [cell.value for cell in row]
        if all(item == onerow[0] for item in onerow):
            empty_rows.append(row[0].row)
    return empty_rows


def list_nan_worksheet_rows(
    *,
    ws: Worksheet,
    min_row: int
) -> list[int]:
    """
    Create list of row numbers of blank worksheet rows.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    min_row : int
        Start row for iteration.

    Returns
    -------
    blank_rows : list[int]
        List of row numbers.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> sheetname = "sheetname"
    >>> ws = wb[sheetname] # doctest: +SKIP
    >>> blank_rows = dd.list_nan_worksheet_rows(
    ...     ws=ws,
    ...     min_row=2
    ... )
    """
    blank_rows = []
    for row in ws.iter_rows(min_row=min_row):
        onerow = [cell.value for cell in row]
        if all(item != item for item in onerow):
            blank_rows.append(row[0].row)
    return blank_rows


def list_rows_with_content(
    *,
    ws: Worksheet,
    min_row: int,
    column: int,
    text: str
) -> list[int]:
    """
    List rows that contain specific text in a specified column.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    min_row : int
        Start row for iteration.
    column : int
        The column to search.
    text : str
        The text to search.

    Returns
    -------
    list[int]
        A list of row numbers.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> rows_with_text = dd.list_rows_with_content(
    ...     ws=ws,
    ...     min_row=2,
    ...     column=11,
    ...     text='ETA'
    ... )
    """
    rows_with_text = []
    for row in ws.iter_rows(
        min_row=min_row,
        min_col=column,
        max_col=column
    ):
        for cell in row:
            if cell.value == text:
                rows_with_text.append(row[0].row)
    return rows_with_text


def read_workbook(
    *,
    filename: Path | str,
    data_only: bool = True
) -> tuple[Workbook, list[str]]:
    """
    Read a workbook, print the Path, and print the sheet names.

    Parameters
    ----------
    filename : Path | str
        The file containing the workbook.
    data_only : bool = True
        If True, read values stored in the cells. If False, read formulae
        stored in the cells.

    Returns
    -------
    tuple[Workbook, list[str]]
        A tuple of a Workbook and a list of sheet names.

            - wb : Workbook
                A workbook.
            - sheet_names : list[str]
                The sheet names in the workbook.

    Examples
    --------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> filename = "../tests/excel_file.xlsx"
    >>> wb, sheet_names = dd.read_workbook(
    ...     filename=filename,
    ...     data_only=True
    ... ) # doctest: +SKIP
    """
    wb = load_workbook(
        filename=filename,
        data_only=data_only
    )
    sheet_names = wb.sheetnames
    return (wb, sheet_names)


def remove_empty_worksheet_rows(
    *,
    ws: Worksheet,
    empty_rows: list[int]
) -> Worksheet:
    """
    Delete empty worksheet rows.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    empty_rows : list[int]
        List of row numbers.

    Returns
    -------
    ws : Worksheet
        A worksheet from a workbook.

    Example
    -------
    Remove empty rows found.
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.remove_empty_worksheet_rows(
    ...     ws=ws,
    ...     empty_rows=[5, 6, 7]
    ... )
    """
    for row_idx in reversed(empty_rows):
        ws.delete_rows(
            idx=row_idx,
            amount=1
        )
    return ws


def remove_worksheet_columns(
    *,
    ws: Worksheet,
    starting_column: int,
    number_of_columns: int
) -> Worksheet:
    """
    Remove worksheet columns.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    starting_column : int
        The first column to remove.
    number_of_columns : int
        The number of columns to remove.

    Returns
    -------
    ws : Worksheet
        A worksheet from a workbook.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.remove_worksheet_columns(
    ...     ws=ws,
    ...     starting_column=13,
    ...     number_of_columns=3
    ... )
    """
    ws.delete_cols(
        idx=starting_column,
        amount=number_of_columns
    )
    return ws


def remove_worksheet_rows(
    *,
    ws: Worksheet,
    rows_to_remove: list[int]
) -> Worksheet:
    """
    Remove worksheet rows.

    Parameters
    ----------
    ws : Worksheet
        A worksheet from a workbook.
    rows_to_remove: list[int]
        The list of row numbers to remove.

    Returns
    -------
    ws : Worksheet
        A worksheet from a workbook.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.remove_worksheet_rows(
    ...     ws=ws,
    ...     rows_to_remove=[4, 5, 6]
    ... )
    """
    for row in reversed(rows_to_remove):
        ws.delete_rows(
            idx=row,
            amount=1
        )
    return ws


def replace_text(
    *,
    ws: Worksheet,
    column: int,
    text: tuple[tuple[str, str]]
) -> Worksheet:
    """
    Search and replace text in a cell.

    Parameters
    ----------
    ws : Worksheet
        The worksheet in which to search and replace text.
    column : int
        The column number in which to search and replace text.
    text ; tuple[tuple[str, str]]
        The search and replace text.

    Returns
    -------
    ws : Worksheet
        The worksheet in which text was replaced.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.replace_text(
    ...     ws=ws,
    ...     column=13,
    ...     text=['old_text', 'new_text']
    ... ) # doctest: +SKIP
    """
    for row in ws.iter_rows(
        min_col=column,
        max_col=column
    ):
        for item in text:
            for cell in row:
                cell.value = str(cell.value).replace(*item)
    return ws


def validate_column_labels(
    *,
    ws: Worksheet,
    column_labels: list[str],
    first_column: int,
    last_column: int,
    row_of_labels: int,
    start_time: float = None,
    stop_time: float = None,
    original_stdout: IO[str] = None,
    output_url: str = None

) -> Worksheet:
    """
    Validate the labels of a worksheet with a desired list of labels

    Parameters
    ----------
    ws : Worksheet
        The worksheet to analyze.
    column_labels : list[str]
        The list of desired column labels.
    first_column : int
        The first column of the label range in the worksheet.
    last_column : int
        The last column of the label range in the worksheet.
    row_of_labels : int
        The row number of the labels in the worksheet.
    start_time : float = None
        The start time of the script.
    stop_time : float = None
        The stop time of the script.
    original_stdout : IO[str] = None
        The original stdout.
    output_url : str = None
        The output url.

    Returns
    -------
    ws : Worksheet,
        The worksheet to analyze.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> ws = dd.validate_column_labels(
    ...     ws=ws,
    ...     column_labels=column_labels,
    ...     first_column=first_column,
    ...     last_column=last_column,
    ...     row_of_labels=row_of_labels,
    ...     start_time=start_time,
    ...     stop_time=time.time(),
    ...     original_stdout=original_stdout,
    ...     output_url=output_url
    ... ) # doctest: +SKIP
    """
    labels_found = []
    for col in range(first_column, last_column + 1):
        labels_found.append(ws.cell(row=row_of_labels, column=col).value)
    if len(labels_found) == len(column_labels) and \
            labels_found[-1] == column_labels[-1]:
        for col, label in zip(
            range(first_column, last_column + 1),
            column_labels
        ):
            ws.cell(
                row=row_of_labels,
                column=col
            ).value = label
    elif stop_time:
        print(
            'Column labels or number of columns are incorrect. '
            'Fix. Re-run script.'
        )
        print('Labels found:')
        print(labels_found)
        print('Length of labels found:', len(labels_found))
        print('Labels should be:')
        print(column_labels)
        print('Length of labels should be:', len(column_labels))
        print('XXX File NOT OK XXX')
        stop_time = stop_time
        report_summary(
            start_time=start_time,
            stop_time=stop_time
        )
        exit_script(
            original_stdout=original_stdout,
            output_url=output_url
        )
    return ws


def unique_list_items(
    *,
    ws: Worksheet,
    row_of_labels: int,
    row_below_labels: int,
    column_name_varname: str,
    text_to_replace: list[str],
    text_to_remove: list[str]
) -> tuple[list[int], list[int]]:
    """
    Determine list of unique items in varname.
    Replace text.

    TODO:
    This function does too many things. Break it up.
    Add detail to docstring.

    Parameters
    ----------
    ws : Worksheet
        The worksheet to analyze.
    row_of_labels : int
        The row number of the labels in the worksheet.
    row_below_labels : int
        The row number below the label row.
    column_name_varname : str
        The label of a column.
    text_to_replace : list[str]
        A list of the text to search for.
    text_to_remove : list[str]
        A list of the text to remove.

    Returns
    -------
    tuple[list[int], list[int]]
        A tuple of a list of unique items and a list of column numbers.

            - varname : tuple[list[int]
                The list of unique items.
            - column_numbers : list[int]]
                The list of column numbers.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> list_part_numbers, list_column_numbers = dd.unique_list_items(
    ...     ws=ws,
    ...     row_of_labels=row_of_labels,
    ...     row_below_labels=row_below_labels,
    ...     column_name_varname=column_name,
    ...     text_to_replace=text_to_replace,
    ...     text_to_remove=text_to_remove
    ... ) # doctest: +SKIP
    """
    column_numbers = [
        col for col in range(ws.min_column, ws.max_column + 1)
    ]
    column_labels = []
    for row in ws.iter_rows(
        min_row=row_of_labels,
        max_row=row_of_labels,
        min_col=column_numbers[0],
        max_col=column_numbers[-1]
    ):
        for cell in row:
            column_labels.append(cell.value)
    column_names_numbers = dict(zip(column_labels, column_numbers))
    varname_replace = []
    for row in ws.iter_rows(
        min_row=row_below_labels,
        min_col=column_names_numbers[column_name_varname],
        max_col=column_names_numbers[column_name_varname]
    ):
        for cell in row:
            if cell.value:
                varname_replace.append(cell.value)
    varname_remove = []
    for item in varname_replace:
        for thing in text_to_replace:
            item = str(item).replace(thing, '')
        varname_remove.append(item)
    varname = [x for x in varname_remove if x not in text_to_remove]
    varname = (list(set(varname)))
    return (varname, column_numbers)


def validate_sheet_names(
    *,
    wb: Workbook,
    filename: Path | str,
    sheet_name: str,
    sheet_names: list[str],
    start_time: float,
    original_stdout: io.TextIOWrapper,
    output_url: str
) -> Workbook:
    """
    Parameters
    ----------
    wb : Workbook
        A workbook.
    filename : Path | str
        The file containing the workbook.
    sheet_name : str
        A sheet name in the workbook.
    sheet_names : list[str]
        The sheet names in the workbook.
    start_time : float
        The start time of the script.
    original_stdout : io.TextIOWrapper
        The buffered text stream for the html output.
    output_url : str
        The html filename.

    Returns
    -------
    wb : Workbook
        The workbook with a revised sheetname.

    Example
    -------
    >>> import dawgdad as dd
    >>> start_time = time.perf_counter()
    >>> stop_time = time.perf_counter()
    >>> filename = "../tests/excel_file.xlsx"
    >>> sheet_name = "sheet_01"
    >>> sheet_names = ["sheet_01", "sheet_02"]
    >>> output_url = "../tests/my_html_file.html"
    >>> original_stdout = dd.html_begin(output_url=output_url)
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> wb = validate_sheet_names(
    ...     wb=wb,
    ...     filename=filename,
    ...     sheet_name=sheet_name,
    ...     sheet_names=sheet_names,
    ...     start_time=start_time,
    ...     original_stdout=original_stdout,
    ...     output_url=output_url
    ... ) # doctest: +SKIP
    """
    if sheet_name not in sheet_names and len(sheet_names) != 1:
        print('Manually rename one of these sheets:')
        print(wb.sheetnames)
        print('Then re-run script')
        print('XXX File NOT OK XXX')
        stop_time = time.time()
        report_summary(
            start_time=start_time,
            stop_time=stop_time
        )
        exit_script(
            original_stdout=original_stdout,
            output_url=output_url
        )
    elif sheet_name not in sheet_names and len(sheet_names) == 1:
        print('One sheet found and it was re-named.')
        ws = wb.active
        ws.title = sheet_name
        wb.save(filename=filename)
    elif sheet_name in sheet_names and len(sheet_names) != 1:
        sheet_names_removed = [x for x in sheet_names if x != sheet_name]
        for sheet in sheet_names_removed:
            wb.remove(worksheet=wb[sheet])
        print('Sheet names removed:')
        print(sheet_names_removed)
        wb.save(filename=filename)
    return wb


def write_dataframe_to_worksheet(
    *,
    ws: Worksheet,
    df: pd.DataFrame,
    index: bool = False,
    header: bool = True
) -> Worksheet:
    """
    Write a dataframe to a worksheet.

    Parameters
    ----------
    ws : Worksheet
        The worksheet to which the dataframe will be written.
    df : pd.DataFrame
        The dataframe of data.
    index : bool = False
        Boolean to determine if dataframe index is written to worksheet.
    header : bool = True
        Boolean to determine if dataframe header is written to worksheet.

    Returns
    -------
    ws : Worksheet
        The worksheet created.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> df = pd.DataFrame(data={
    ...     'X': [25.0, 24.0, 35.5, np.nan, 23.1],
    ...     'Y': [27, 24, np.nan, 23, np.nan],
    ...     'Z': ['a', 'b', np.nan, 'd', 'e']
    ... })
    >>> ws = dd.write_dataframe_to_worksheet(
    ...     ws=ws,
    ...     df=df,
    ...     index=False,
    ...     header=True
    ... )
    """
    for row in dataframe_to_rows(
        df=df,
        index=index,
        header=header
    ):
        ws.append(row)
    return ws


def number_non_empty_rows(
    *,
    ws: Worksheet,
    column_number: int,
    start_row: int,
) -> int:
    """
    Determine the number of non-empty rows for a single column.

    Parameters
    ----------
    ws : Worksheet
        The worksheet to analyze.
    column_number : int
        The desired column number.
    start_row : int
        The row at which to start evaluating cells.

    Returns
    -------
    row_count : int
        The number of non-empty rows.

    Example
    -------
    >>> import dawgdad as dd
    >>> wb = Workbook()
    >>> ws = wb.active
    >>> start_row = 2
    >>> column_number = 1
    >>> row_count = dd.number_non_empty_rows(
    ...     ws=ws,
    ...     column_number=column_number,
    ...     start_row=start_row,
    ... )
    """
    row_count = 0
    for row in ws.iter_rows(
        min_row=start_row,
        min_col=column_number,
        max_col=column_number
    ):
        for cell in row:
            if cell.value:
                row_count += 1
    return row_count


__all__ = (
    'list_empty_except_nan_worksheet_rows',
    'list_empty_and_nan_worksheet_rows',
    'list_duplicate_worksheet_rows',
    'change_case_worksheet_columns',
    'write_dataframe_to_worksheet',
    'remove_empty_worksheet_rows',
    'remove_worksheet_columns',
    'list_nan_worksheet_rows',
    'list_rows_with_content',
    'validate_column_labels',
    'number_non_empty_rows',
    'remove_worksheet_rows',
    'validate_sheet_names',
    'autofit_column_width',
    'unique_list_items',
    'cell_fill_down',
    'read_workbook',
    'replace_text',
    'exit_script',
    'cell_style',
)
